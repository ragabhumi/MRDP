# -*- coding: utf-8 -*-
"""
Created on Thu Oct 06 20:57:44 2016

@author: YOSI
"""

from datetime import datetime
import time
import numpy as np
import re
import string
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
import errno
import glob
from calendar import monthrange

def make_sure_path_exist(path):
    try:
        os.makedirs(path)
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise
def save_excel(IAGA_folder,IMFV_folder):	
    #Import data IAGA-2002   
    pathIAGA = str(IAGA_folder)
    pathIMFV = str(IMFV_folder)
    Bx = list(np.tile(np.nan,1440))
    By = list(np.tile(np.nan,1440))
    Bz = list(np.tile(np.nan,1440))
    Bf = list(np.tile(np.nan,1440))
    Bh = list(np.tile(np.nan,1440))
    Bd = list(np.tile(np.nan,1440))
    Bi = list(np.tile(np.nan,1440))
    
    first_data = os.path.basename(glob.glob(os.path.join(pathIAGA, '*.min'))[0])
    tahun1 = int(first_data[3:7])
    bulan1 = int(first_data[7:9])
    tanggal1 = 1
    date1 = datetime(year=tahun1,month=bulan1,day=tanggal1)
    hari_bulan = monthrange(tahun1, bulan1)[1]

    x_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    y_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    z_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    f_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    h_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    d_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
    i_mean=[['AM' for x in range(24)] for x in range(hari_bulan)] 
             
    x_mean_hour_month1=[np.nan for x in range(24)]
    y_mean_hour_month1=[np.nan for x in range(24)]
    z_mean_hour_month1=[np.nan for x in range(24)]
    f_mean_hour_month1=[np.nan for x in range(24)]
    h_mean_hour_month1=[np.nan for x in range(24)]
    d_mean_hour_month1=[np.nan for x in range(24)]
    i_mean_hour_month1=[np.nan for x in range(24)]
                        
    x_mean_hour_month=[np.nan for x in range(24)]
    y_mean_hour_month=[np.nan for x in range(24)]
    z_mean_hour_month=[np.nan for x in range(24)]
    f_mean_hour_month=[np.nan for x in range(24)]
    h_mean_hour_month=[np.nan for x in range(24)]
    d_mean_hour_month=[np.nan for x in range(24)]
    i_mean_hour_month=[np.nan for x in range(24)]

             
    x_mean_day=list(np.tile(np.nan,hari_bulan))
    y_mean_day=list(np.tile(np.nan,hari_bulan))
    z_mean_day=list(np.tile(np.nan,hari_bulan))
    f_mean_day=list(np.tile(np.nan,hari_bulan))
    h_mean_day=list(np.tile(np.nan,hari_bulan))
    d_mean_day=list(np.tile(np.nan,hari_bulan))
    i_mean_day=list(np.tile(np.nan,hari_bulan))
             
    k_i = ['AM']*hari_bulan*8
    a_i = [-1]*hari_bulan*8
    sk = ['']*hari_bulan
    A_i = [-1]*hari_bulan
    status = ['']*hari_bulan
    m = 0

    for filename in glob.glob(os.path.join(pathIAGA, '*.min')):
        
        try:
            with open(filename) as f_lemi:
                for num, line in enumerate(f_lemi, 1):
                    if 'DATE' in line:
                        skip_line = num
            with open(filename) as f_lemi:
                print os.path.basename(filename)
                content = f_lemi.readlines()[skip_line:]
                j = 0
                for i in range(0,len(content)):
                    data = re.split('\s+',content[i])
                    if data[3]!='99999.00':
                        Bh[i] = float(data[3])
                    else:
                        Bh[i] = np.nan
                    if data[4]!='99999.00':
                        Bd[i] = float(data[4])
                    else:
                        Bd[i] = np.nan
                    if data[5]!='99999.00':
                        Bz[i] = float(data[5])
                    else:
                        Bz[i] = np.nan
                    if data[6]!='99999.00':
                        Bf[i] = float(data[6])
                    else:
                        Bf[i] = np.nan
                    Bx[i] = Bh[i]*np.cos(np.deg2rad(Bd[i]/60))
                    By[i] = Bh[i]*np.sin(np.deg2rad(Bd[i]/60))
                    Bi[i] = np.rad2deg(np.arctan(Bz[i]/Bh[i]))
                    j = j + 1
        except IOError:
            continue
        
    #Hitung rata-rata satu jam
        for k in range(0,24):
            if np.count_nonzero(np.isnan(Bx[(0+(60*k)):((60+(60*k)))])) <= 6:
                x_mean[m][k]='%5.0f' %np.nanmean(Bx[(0+(60*k)):((60+(60*k)))])
            else:
                x_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(By[(0+(60*k)):((60+(60*k)))])) <= 6:
                y_mean[m][k]='%6.1f' %np.nanmean(By[(0+(60*k)):((60+(60*k)))])
            else:
                y_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(Bz[(0+(60*k)):((60+(60*k)))])) <= 6:
                z_mean[m][k]='%5.0f' %np.nanmean(Bz[(0+(60*k)):((60+(60*k)))])
            else:
                z_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(Bf[(0+(60*k)):((60+(60*k)))])) <= 6:
                f_mean[m][k]='%5.0f' %np.nanmean(Bf[(0+(60*k)):((60+(60*k)))])
            else:
                f_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(Bh[(0+(60*k)):((60+(60*k)))])) <= 6:
                h_mean[m][k]='%5.0f' %np.nanmean(Bh[(0+(60*k)):((60+(60*k)))])
            else:
                h_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(Bd[(0+(60*k)):((60+(60*k)))])) <= 6:
                d_mean[m][k]='%6.2f' %np.nanmean(Bd[(0+(60*k)):((60+(60*k)))])
            else:
                d_mean[m][k]='AM'
            if np.count_nonzero(np.isnan(Bi[(0+(60*k)):((60+(60*k)))])) <= 6:
                i_mean[m][k]='%6.2f' %np.nanmean(Bi[(0+(60*k)):((60+(60*k)))])
            else:
                i_mean[m][k]='AM'
        m = m+1
    
    # Hitung rata-rata per hari
    for n in range(0,hari_bulan):
        if 'AM' in x_mean[n]:
            x_mean_day[n] = ''
        else:
            x_mean_day[n] = '%5.0f' %np.mean([float(i) for i in x_mean[n]])
        if 'AM' in y_mean[n]:
            y_mean_day[n] = ''
        else:
            y_mean_day[n] = '%6.1f' %np.mean([float(i) for i in y_mean[n]])
        if 'AM' in x_mean[n]:
            z_mean_day[n] = ''
        else:
            z_mean_day[n] = '%5.0f' %np.mean([float(i) for i in z_mean[n]])
        if 'AM' in x_mean[n]:
            f_mean_day[n] = ''
        else:
            f_mean_day[n] = '%5.0f' %np.mean([float(i) for i in f_mean[n]])
        if 'AM' in x_mean[n]:
            h_mean_day[n] = ''
        else:
            h_mean_day[n] = '%5.0f' %np.mean([float(i) for i in h_mean[n]])
        if 'AM' in x_mean[n]:
            d_mean_day[n] = ''
        else:
            d_mean_day[n] = '%6.2f' %np.mean([float(i) for i in d_mean[n]])
        if 'AM' in x_mean[n]:
            i_mean_day[n] = ''
        else:
            i_mean_day[n] = '%6.2f' %np.mean([float(i) for i in i_mean[n]])
            
    # Hitung rata-rata per jam per bulan
    for p in range(0,24):
        x_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(x_mean))[p]]
        y_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(y_mean))[p]]
        z_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(z_mean))[p]]
        f_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(f_mean))[p]]
        h_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(h_mean))[p]]
        d_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(d_mean))[p]]
        i_mean_hour_month1[p] = [w.replace('AM', str(np.nan)) for w in (np.transpose(i_mean))[p]]

        x_mean_hour_month[p] = '%5.0f' %np.nanmean([float(i) for i in x_mean_hour_month1[p]])
        y_mean_hour_month[p] = '%6.1f' %np.nanmean([float(i) for i in y_mean_hour_month1[p]])
        z_mean_hour_month[p] = '%5.0f' %np.nanmean([float(i) for i in z_mean_hour_month1[p]])
        f_mean_hour_month[p] = '%5.0f' %np.nanmean([float(i) for i in f_mean_hour_month1[p]])
        h_mean_hour_month[p] = '%5.0f' %np.nanmean([float(i) for i in h_mean_hour_month1[p]])
        d_mean_hour_month[p] = '%6.2f' %np.nanmean([float(i) for i in d_mean_hour_month1[p]])
        i_mean_hour_month[p] = '%6.2f' %np.nanmean([float(i) for i in i_mean_hour_month1[p]])
    
    # Hitung rata-rata per bulan
    x_month = '%5.0f' %(sum([float(i) for i in x_mean_hour_month])/24)
    y_month = '%6.1f' %(sum([float(i) for i in y_mean_hour_month])/24)
    z_month = '%5.0f' %(sum([float(i) for i in z_mean_hour_month])/24)
    f_month = '%5.0f' %(sum([float(i) for i in f_mean_hour_month])/24)
    h_month = '%5.0f' %(sum([float(i) for i in h_mean_hour_month])/24)
    d_month = '%6.2f' %(sum([float(i) for i in d_mean_hour_month])/24)
    i_month = '%6.2f' %(sum([float(i) for i in i_mean_hour_month])/24)
    
    #Hitung K-Index sebulan
    os.system('kasm TUN:01%s:%2.0f 300 xy %s\\data %s\\' % (string.upper(date1.strftime("%b%Y")),hari_bulan,pathIMFV,pathIMFV))

    #Proses pemilahan data
    with open(pathIMFV + '\\data.dka') as f:
        content = f.readlines()
        for j in range(6,len(content)):
            data_k = re.split(' ',content[j])
            for i in range(0,len(data_k)-14):
                data_k.remove('')
            index_kk = time.strptime(data_k[0],'%d-%b-%y').tm_mday-1
            sk[index_kk] = int(data_k[10])
            if sk[index_kk] == -1:
                sk[index_kk] = ''
            # print sk
            for m in range(0,8):
                k_i[m+((index_kk)*8)] = int(data_k[m+2])
                if (k_i[m+((index_kk)*8)]==0):
                    a_i[m+((index_kk)*8)]=0
                elif (k_i[m+((index_kk)*8)]==1):
                    a_i[m+((index_kk)*8)]=3
                elif (k_i[m+((index_kk)*8)]==2):
                    a_i[m+((index_kk)*8)]=6	
                elif (k_i[m+((index_kk)*8)]==3):
                    a_i[m+((index_kk)*8)]=12	
                elif (k_i[m+((index_kk)*8)]==4):
                    a_i[m+((index_kk)*8)]=24	
                elif (k_i[m+((index_kk)*8)]==5):
                    a_i[m+((index_kk)*8)]=40	
                elif (k_i[m+((index_kk)*8)]==6):
                    a_i[m+((index_kk)*8)]=70
                elif (k_i[m+((index_kk)*8)]==7):
                    a_i[m+((index_kk)*8)]=120	    
                elif (k_i[m+((index_kk)*8)]==8):
                    a_i[m+((index_kk)*8)]=200	    
                elif (k_i[m+((index_kk)*8)]==9):
                    a_i[m+((index_kk)*8)]=300
                elif (k_i[m+((index_kk)*8)]==-1):
                    a_i[m+((index_kk)*8)]=-1
                    k_i[m+((index_kk)*8)] = 'AM'
    

    #Proses penentuan Character Magnetic Activity
    for n in range(0,hari_bulan):
        if sk[n]=='':
            status[n]=''
        else:
            A_i[n] = sum(a_i[0+(n*8):8+(n*8)])/8
            if A_i<0:
                status[n]=''
            elif (A_i[n]>=0 and A_i[n]<=30):
                status[n]='Hari Tenang'
            elif (A_i[n]>30 and A_i[n]<=50):
                status[n]='Badai Lemah'
            elif (A_i[n]>50 and A_i[n]<=100):
                status[n]='Badai Menengah'
            elif (A_i[n]>100):
                status[n]='Badai Kuat'

        
    #Output file format excel
    wb = Workbook()
    make_sure_path_exist(pathIAGA)
    fileout = pathIAGA + '\\' + string.upper(date1.strftime("%b%Y")) + '.xlsx'

    # KOMPONEN X
    ws1 = wb.active
    ws1.merge_cells('A2:Z2');ws1.merge_cells('A3:E3');ws1.merge_cells('Y3:Z3')
    ws1.merge_cells('B4:B5');ws1.merge_cells('C4:C5');ws1.merge_cells('D4:D5')
    ws1.merge_cells('E4:E5');ws1.merge_cells('F4:F5');ws1.merge_cells('G4:G5')
    ws1.merge_cells('H4:H5');ws1.merge_cells('I4:I5');ws1.merge_cells('J4:J5')
    ws1.merge_cells('K4:K5');ws1.merge_cells('L4:L5');ws1.merge_cells('M4:M5')
    ws1.merge_cells('N4:N5');ws1.merge_cells('O4:O5');ws1.merge_cells('P4:P5')
    ws1.merge_cells('Q4:Q5');ws1.merge_cells('R4:R5');ws1.merge_cells('S4:S5')
    ws1.merge_cells('T4:T5');ws1.merge_cells('U4:U5');ws1.merge_cells('V4:V5')
    ws1.merge_cells('W4:W5');ws1.merge_cells('X4:X5');ws1.merge_cells('Y4:Y5')
    ws1.merge_cells('Z4:Z5')
    ws1['A2'] = 'Hourly Mean Values of X in nT From Digital Magnet'
    ws1['A2'].alignment = Alignment(horizontal="center")
    ws1['A3'] = 'Tuntungan Magnetic Observatory'
    ws1['A4'] = 'Hour'
    ws1['A5'] = 'Day'
    ws1['Y3'] = date1.strftime("%B %Y")
    ws1['B4'] = '1';ws1['C4'] = '2';ws1['D4'] = '3';ws1['E4'] = '4';ws1['F4'] = '5'
    ws1['G4'] = '6';ws1['H4'] = '7';ws1['I4'] = '8';ws1['J4'] = '9';ws1['K4'] = '10'
    ws1['L4'] = '11';ws1['M4'] = '12';ws1['N4'] = '13';ws1['O4'] = '14';ws1['P4'] = '15'
    ws1['Q4'] = '16';ws1['R4'] = '17';ws1['S4'] = '18';ws1['T4'] = '19';ws1['U4'] = '20'
    ws1['V4'] = '21';ws1['W4'] = '22';ws1['X4'] = '23';ws1['Y4'] = '24'
    ws1['Z4'] = 'Mean'; ws1['A37'] = 'Mean'
    ws1['B38'] = 'Remark'; ws1['B39'] = 'AM = Alat Mati'
    ws1.title = "X"

    # KOMPONEN Y
    ws2 = wb.create_sheet()
    ws2.merge_cells('A2:Z2');ws2.merge_cells('A3:E3');ws2.merge_cells('Y3:Z3')
    ws2.merge_cells('B4:B5');ws2.merge_cells('C4:C5');ws2.merge_cells('D4:D5')
    ws2.merge_cells('E4:E5');ws2.merge_cells('F4:F5');ws2.merge_cells('G4:G5')
    ws2.merge_cells('H4:H5');ws2.merge_cells('I4:I5');ws2.merge_cells('J4:J5')
    ws2.merge_cells('K4:K5');ws2.merge_cells('L4:L5');ws2.merge_cells('M4:M5')
    ws2.merge_cells('N4:N5');ws2.merge_cells('O4:O5');ws2.merge_cells('P4:P5')
    ws2.merge_cells('Q4:Q5');ws2.merge_cells('R4:R5');ws2.merge_cells('S4:S5')
    ws2.merge_cells('T4:T5');ws2.merge_cells('U4:U5');ws2.merge_cells('V4:V5')
    ws2.merge_cells('W4:W5');ws2.merge_cells('X4:X5');ws2.merge_cells('Y4:Y5')
    ws2.merge_cells('Z4:Z5')
    ws2['A2'] = 'Hourly Mean Values of Y in nT From Digital Magnet'
    ws2['A2'].alignment = Alignment(horizontal="center")
    ws2['A3'] = 'Tuntungan Magnetic Observatory'
    ws2['A4'] = 'Hour'
    ws2['A5'] = 'Day'
    ws2['Y3'] = date1.strftime("%B %Y")
    ws2['B4'] = '1';ws2['C4'] = '2';ws2['D4'] = '3';ws2['E4'] = '4';ws2['F4'] = '5'
    ws2['G4'] = '6';ws2['H4'] = '7';ws2['I4'] = '8';ws2['J4'] = '9';ws2['K4'] = '10'
    ws2['L4'] = '11';ws2['M4'] = '12';ws2['N4'] = '13';ws2['O4'] = '14';ws2['P4'] = '15'
    ws2['Q4'] = '16';ws2['R4'] = '17';ws2['S4'] = '18';ws2['T4'] = '19';ws2['U4'] = '20'
    ws2['V4'] = '21';ws2['W4'] = '22';ws2['X4'] = '23';ws2['Y4'] = '24'
    ws2['Z4'] = 'Mean'; ws2['A37'] = 'Mean'
    ws2['B38'] = 'Remark'; ws2['B39'] = 'AM = Alat Mati'
    ws2.title = "Y"

    # KOMPONEN Z
    ws3 = wb.create_sheet()
    ws3.merge_cells('A2:Z2');ws3.merge_cells('A3:E3');ws3.merge_cells('Y3:Z3')
    ws3.merge_cells('B4:B5');ws3.merge_cells('C4:C5');ws3.merge_cells('D4:D5')
    ws3.merge_cells('E4:E5');ws3.merge_cells('F4:F5');ws3.merge_cells('G4:G5')
    ws3.merge_cells('H4:H5');ws3.merge_cells('I4:I5');ws3.merge_cells('J4:J5')
    ws3.merge_cells('K4:K5');ws3.merge_cells('L4:L5');ws3.merge_cells('M4:M5')
    ws3.merge_cells('N4:N5');ws3.merge_cells('O4:O5');ws3.merge_cells('P4:P5')
    ws3.merge_cells('Q4:Q5');ws3.merge_cells('R4:R5');ws3.merge_cells('S4:S5')
    ws3.merge_cells('T4:T5');ws3.merge_cells('U4:U5');ws3.merge_cells('V4:V5')
    ws3.merge_cells('W4:W5');ws3.merge_cells('X4:X5');ws3.merge_cells('Y4:Y5')
    ws3.merge_cells('Z4:Z5')
    ws3['A2'] = 'Hourly Mean Values of Z in nT From Digital Magnet'
    ws3['A2'].alignment = Alignment(horizontal="center")
    ws3['A3'] = 'Tuntungan Magnetic Observatory'
    ws3['A4'] = 'Hour'
    ws3['A5'] = 'Day'
    ws3['Y3'] = date1.strftime("%B %Y")
    ws3['B4'] = '1';ws3['C4'] = '2';ws3['D4'] = '3';ws3['E4'] = '4';ws3['F4'] = '5'
    ws3['G4'] = '6';ws3['H4'] = '7';ws3['I4'] = '8';ws3['J4'] = '9';ws3['K4'] = '10'
    ws3['L4'] = '11';ws3['M4'] = '12';ws3['N4'] = '13';ws3['O4'] = '14';ws3['P4'] = '15'
    ws3['Q4'] = '16';ws3['R4'] = '17';ws3['S4'] = '18';ws3['T4'] = '19';ws3['U4'] = '20'
    ws3['V4'] = '21';ws3['W4'] = '22';ws3['X4'] = '23';ws3['Y4'] = '24'
    ws3['Z4'] = 'Mean'; ws3['A37'] = 'Mean'
    ws3['B38'] = 'Remark'; ws3['B39'] = 'AM = Alat Mati'
    ws3.title = "Z"
    
    # KOMPONEN H
    ws4 = wb.create_sheet()
    ws4.merge_cells('A2:Z2');ws4.merge_cells('A3:E3');ws4.merge_cells('Y3:Z3')
    ws4.merge_cells('B4:B5');ws4.merge_cells('C4:C5');ws4.merge_cells('D4:D5')
    ws4.merge_cells('E4:E5');ws4.merge_cells('F4:F5');ws4.merge_cells('G4:G5')
    ws4.merge_cells('H4:H5');ws4.merge_cells('I4:I5');ws4.merge_cells('J4:J5')
    ws4.merge_cells('K4:K5');ws4.merge_cells('L4:L5');ws4.merge_cells('M4:M5')
    ws4.merge_cells('N4:N5');ws4.merge_cells('O4:O5');ws4.merge_cells('P4:P5')
    ws4.merge_cells('Q4:Q5');ws4.merge_cells('R4:R5');ws4.merge_cells('S4:S5')
    ws4.merge_cells('T4:T5');ws4.merge_cells('U4:U5');ws4.merge_cells('V4:V5')
    ws4.merge_cells('W4:W5');ws4.merge_cells('X4:X5');ws4.merge_cells('Y4:Y5')
    ws4.merge_cells('Z4:Z5')
    ws4['A2'] = 'Hourly Mean Values of H in nT From Digital Magnet'
    ws4['A2'].alignment = Alignment(horizontal="center")
    ws4['A3'] = 'Tuntungan Magnetic Observatory'
    ws4['A4'] = 'Hour'
    ws4['A5'] = 'Day'
    ws4['Y3'] = date1.strftime("%B %Y")
    ws4['B4'] = '1';ws4['C4'] = '2';ws4['D4'] = '3';ws4['E4'] = '4';ws4['F4'] = '5'
    ws4['G4'] = '6';ws4['H4'] = '7';ws4['I4'] = '8';ws4['J4'] = '9';ws4['K4'] = '10'
    ws4['L4'] = '11';ws4['M4'] = '12';ws4['N4'] = '13';ws4['O4'] = '14';ws4['P4'] = '15'
    ws4['Q4'] = '16';ws4['R4'] = '17';ws4['S4'] = '18';ws4['T4'] = '19';ws4['U4'] = '20'
    ws4['V4'] = '21';ws4['W4'] = '22';ws4['X4'] = '23';ws4['Y4'] = '24'
    ws4['Z4'] = 'Mean'; ws4['A37'] = 'Mean'
    ws4['B38'] = 'Remark'; ws4['B39'] = 'AM = Alat Mati'
    ws4.title = "H"
  
    # KOMPONEN F
    ws5 = wb.create_sheet()
    ws5.merge_cells('A2:Z2');ws5.merge_cells('A3:E3');ws5.merge_cells('Y3:Z3')
    ws5.merge_cells('B4:B5');ws5.merge_cells('C4:C5');ws5.merge_cells('D4:D5')
    ws5.merge_cells('E4:E5');ws5.merge_cells('F4:F5');ws5.merge_cells('G4:G5')
    ws5.merge_cells('H4:H5');ws5.merge_cells('I4:I5');ws5.merge_cells('J4:J5')
    ws5.merge_cells('K4:K5');ws5.merge_cells('L4:L5');ws5.merge_cells('M4:M5')
    ws5.merge_cells('N4:N5');ws5.merge_cells('O4:O5');ws5.merge_cells('P4:P5')
    ws5.merge_cells('Q4:Q5');ws5.merge_cells('R4:R5');ws5.merge_cells('S4:S5')
    ws5.merge_cells('T4:T5');ws5.merge_cells('U4:U5');ws5.merge_cells('V4:V5')
    ws5.merge_cells('W4:W5');ws5.merge_cells('X4:X5');ws5.merge_cells('Y4:Y5')
    ws5.merge_cells('Z4:Z5')
    ws5['A2'] = 'Hourly Mean Values of F in nT From Digital Magnet'
    ws5['A2'].alignment = Alignment(horizontal="center")
    ws5['A3'] = 'Tuntungan Magnetic Observatory'
    ws5['A4'] = 'Hour'
    ws5['A5'] = 'Day'
    ws5['Y3'] = date1.strftime("%B %Y")
    ws5['B4'] = '1';ws5['C4'] = '2';ws5['D4'] = '3';ws5['E4'] = '4';ws5['F4'] = '5'
    ws5['G4'] = '6';ws5['H4'] = '7';ws5['I4'] = '8';ws5['J4'] = '9';ws5['K4'] = '10'
    ws5['L4'] = '11';ws5['M4'] = '12';ws5['N4'] = '13';ws5['O4'] = '14';ws5['P4'] = '15'
    ws5['Q4'] = '16';ws5['R4'] = '17';ws5['S4'] = '18';ws5['T4'] = '19';ws5['U4'] = '20'
    ws5['V4'] = '21';ws5['W4'] = '22';ws5['X4'] = '23';ws5['Y4'] = '24'
    ws5['Z4'] = 'Mean'; ws5['A37'] = 'Mean'
    ws5['B38'] = 'Remark'; ws5['B39'] = 'AM = Alat Mati'
    ws5.title = "F"

    # KOMPONEN D
    ws6 = wb.create_sheet()
    ws6.merge_cells('A2:Z2');ws6.merge_cells('A3:E3');ws6.merge_cells('Y3:Z3')
    ws6.merge_cells('B4:B5');ws6.merge_cells('C4:C5');ws6.merge_cells('D4:D5')
    ws6.merge_cells('E4:E5');ws6.merge_cells('F4:F5');ws6.merge_cells('G4:G5')
    ws6.merge_cells('H4:H5');ws6.merge_cells('I4:I5');ws6.merge_cells('J4:J5')
    ws6.merge_cells('K4:K5');ws6.merge_cells('L4:L5');ws6.merge_cells('M4:M5')
    ws6.merge_cells('N4:N5');ws6.merge_cells('O4:O5');ws6.merge_cells('P4:P5')
    ws6.merge_cells('Q4:Q5');ws6.merge_cells('R4:R5');ws6.merge_cells('S4:S5')
    ws6.merge_cells('T4:T5');ws6.merge_cells('U4:U5');ws6.merge_cells('V4:V5')
    ws6.merge_cells('W4:W5');ws6.merge_cells('X4:X5');ws6.merge_cells('Y4:Y5')
    ws6.merge_cells('Z4:Z5')
    ws6['A2'] = 'Hourly Mean Values of D in arcMin From Digital Magnet'
    ws6['A2'].alignment = Alignment(horizontal="center")
    ws6['A3'] = 'Tuntungan Magnetic Observatory'
    ws6['A4'] = 'Hour'
    ws6['A5'] = 'Day'
    ws6['Y3'] = date1.strftime("%B %Y")
    ws6['B4'] = '1';ws6['C4'] = '2';ws6['D4'] = '3';ws6['E4'] = '4';ws6['F4'] = '5'
    ws6['G4'] = '6';ws6['H4'] = '7';ws6['I4'] = '8';ws6['J4'] = '9';ws6['K4'] = '10'
    ws6['L4'] = '11';ws6['M4'] = '12';ws6['N4'] = '13';ws6['O4'] = '14';ws6['P4'] = '15'
    ws6['Q4'] = '16';ws6['R4'] = '17';ws6['S4'] = '18';ws6['T4'] = '19';ws6['U4'] = '20'
    ws6['V4'] = '21';ws6['W4'] = '22';ws6['X4'] = '23';ws6['Y4'] = '24'
    ws6['Z4'] = 'Mean'; ws6['A37'] = 'Mean'
    ws6['B38'] = 'Remark'; ws6['B39'] = 'AM = Alat Mati'
    ws6.title = "D"
 
    # KOMPONEN I
    ws7 = wb.create_sheet()
    ws7.merge_cells('A2:Z2');ws7.merge_cells('A3:E3');ws7.merge_cells('Y3:Z3')
    ws7.merge_cells('B4:B5');ws7.merge_cells('C4:C5');ws7.merge_cells('D4:D5')
    ws7.merge_cells('E4:E5');ws7.merge_cells('F4:F5');ws7.merge_cells('G4:G5')
    ws7.merge_cells('H4:H5');ws7.merge_cells('I4:I5');ws7.merge_cells('J4:J5')
    ws7.merge_cells('K4:K5');ws7.merge_cells('L4:L5');ws7.merge_cells('M4:M5')
    ws7.merge_cells('N4:N5');ws7.merge_cells('O4:O5');ws7.merge_cells('P4:P5')
    ws7.merge_cells('Q4:Q5');ws7.merge_cells('R4:R5');ws7.merge_cells('S4:S5')
    ws7.merge_cells('T4:T5');ws7.merge_cells('U4:U5');ws7.merge_cells('V4:V5')
    ws7.merge_cells('W4:W5');ws7.merge_cells('X4:X5');ws7.merge_cells('Y4:Y5')
    ws7.merge_cells('Z4:Z5')
    ws7['A2'] = 'Hourly Mean Values of I in Degree From Digital Magnet'
    ws7['A2'].alignment = Alignment(horizontal="center")
    ws7['A3'] = 'Tuntungan Magnetic Observatory'
    ws7['A4'] = 'Hour'
    ws7['A5'] = 'Day'
    ws7['Y3'] = date1.strftime("%B %Y")
    ws7['B4'] = '1';ws7['C4'] = '2';ws7['D4'] = '3';ws7['E4'] = '4';ws7['F4'] = '5'
    ws7['G4'] = '6';ws7['H4'] = '7';ws7['I4'] = '8';ws7['J4'] = '9';ws7['K4'] = '10'
    ws7['L4'] = '11';ws7['M4'] = '12';ws7['N4'] = '13';ws7['O4'] = '14';ws7['P4'] = '15'
    ws7['Q4'] = '16';ws7['R4'] = '17';ws7['S4'] = '18';ws7['T4'] = '19';ws7['U4'] = '20'
    ws7['V4'] = '21';ws7['W4'] = '22';ws7['X4'] = '23';ws7['Y4'] = '24'
    ws7['Z4'] = 'Mean'; ws7['A37'] = 'Mean'
    ws7['B38'] = 'Remark'; ws7['B39'] = 'AM = Alat Mati'
    ws7.title = "I"

    # K-INDEX
    ws8 = wb.create_sheet()
    ws8.merge_cells('A1:K1');ws8.merge_cells('A8:K8');ws8.merge_cells('A9:K9')
    ws8.merge_cells('A10:A11');ws8.merge_cells('B10:B11');ws8.merge_cells('C10:C11')
    ws8.merge_cells('D10:D11');ws8.merge_cells('E10:E11');ws8.merge_cells('F10:F11')
    ws8.merge_cells('G10:G11');ws8.merge_cells('H10:H11');ws8.merge_cells('I10:I11')
    ws8.merge_cells('J10:J11');ws8.merge_cells('K10:K11')
    ws8['A1'] = 'M A G N E T I C  A C T I V I T Y'
    ws8['A1'].alignment = Alignment(horizontal="center")
    ws8['A4'] = 'Observatory';ws8['C4'] = ': Tuntungan  - %s  %s'%(string.upper(date1.strftime("%B")),tahun1)
    ws8['A5'] = 'Geog. Latitude'; ws8['C5'] = ': 03 30 01.4 N'; ws8['E5'] = 'Geom. Latitude'
    ws8['I5'] = 'Type of instr  : LEMI-018'
    ws8['A6'] = 'Geog. Long.';ws8['C6'] = ': 98 33 51.6 E';ws8['E6'] = 'Geom. Longitute'
    ws8['A8'] = 'K - Indices for three hours interval'
    ws8['A9'] = 'K - 9 = 300 gammas'
    ws8['A10'] = 'DATE';ws8['B10'] = '00-03';ws8['C10'] = '03-06';ws8['D10'] = '06-09'
    ws8['E10'] = '09-12';ws8['F10'] = '12-15';ws8['G10'] = '15-18';ws8['H10'] = '18-21'
    ws8['I10'] = '21-24';ws8['J10'] = 'SUM';ws8['K10'] = 'CHARACTER'
    ws8.title = "K"
    
    # GEOMAGNETIC STORM
    ws9 = wb.create_sheet()
    ws9.merge_cells('A1:O1');ws9.merge_cells('A2:E2');ws9.merge_cells('L2:O2')
    ws9.merge_cells('J3:O3');ws9.merge_cells('A4:C4');ws9.merge_cells('E4:G4')
    ws9.merge_cells('H4:I4');ws9.merge_cells('J4:L4');ws9.merge_cells('M4:N4')
    ws9.merge_cells('O4:O5');
    ws9['A1'] = 'PRINCIPAL MAGNETIC STORM'
    ws9['A1'].alignment = Alignment(horizontal="center")
    ws9['A2'] = 'Tuntungan Magnetic Observatory';ws9['L2'] = '%s  %s'%(string.upper(date1.strftime("%B")),tahun1);ws9['J3'] = 'UTC Time'
    ws9['A4'] = 'UT Begin'; ws9['D4'] = 'Type'; ws9['E4'] = 'Amplitude'; ws9['H4'] = 'Max. 3 hr Kindices'
    ws9['J4'] = 'Ranges'; ws9['M4'] = 'UT End'; ws9['O4'] = 'Remark'
    ws9['A5'] = 'dd';ws9['B5'] = 'hh';ws9['C5'] = 'mm'
    ws9['E5'] = 'H (nT)'; ws9['F5'] = 'D (Min)'; ws9['G5'] = 'Z (nT)'
    ws9['H5'] = 'day (3 hr period)'; ws9['I5'] = 'K'
    ws9['J5'] = 'H (nT)'; ws9['K5'] = 'D (Min)'; ws9['L5'] = 'Z (nT)'
    ws9['M5'] = 'dd'; ws9['N5'] = 'hh mm'
    ws9.title = "geomag storm"	

    #Save data sebagai excel
    for rows in range(1, hari_bulan+1):
        ws1['A'+str(rows+5)] = rows
        ws1['Z'+str(rows+5)] = x_mean_day[rows-1]
        for columns in range(1,25):
            ws1.cell(column=columns+1,row=rows+5,value=('%s' %x_mean[rows-1][columns-1]))
        ws2['A'+str(rows+5)] = rows
        ws2['Z'+str(rows+5)] = y_mean_day[rows-1]
        for columns in range(1,25):
            ws2.cell(column=columns+1,row=rows+5,value=('%s' %y_mean[rows-1][columns-1]))
        ws3['A'+str(rows+5)] = rows
        ws3['Z'+str(rows+5)] = z_mean_day[rows-1]
        for columns in range(1,25):
            ws3.cell(column=columns+1,row=rows+5,value=('%s' %z_mean[rows-1][columns-1]))
        ws4['A'+str(rows+5)] = rows
        ws4['Z'+str(rows+5)] = h_mean_day[rows-1]
        for columns in range(1,25):
            ws4.cell(column=columns+1,row=rows+5,value=('%s' %h_mean[rows-1][columns-1]))
        ws5['A'+str(rows+5)] = rows
        ws5['Z'+str(rows+5)] = f_mean_day[rows-1]
        for columns in range(1,25):
            ws5.cell(column=columns+1,row=rows+5,value=('%s' %f_mean[rows-1][columns-1]))
        ws6['A'+str(rows+5)] = rows
        ws6['Z'+str(rows+5)] = d_mean_day[rows-1]
        for columns in range(1,25):
            ws6.cell(column=columns+1,row=rows+5,value=('%s' %d_mean[rows-1][columns-1]))
        ws7['A'+str(rows+5)] = rows
        ws7['Z'+str(rows+5)] = i_mean_day[rows-1]
        for columns in range(1,25):
            ws7.cell(column=columns+1,row=rows+5,value=('%s' %i_mean[rows-1][columns-1]))
        ws8['A'+str(rows+11)] = rows
        for columns in range(1,9):
            ws8.cell(column=columns+1,row=rows+11,value=('%s' %k_i[(8*(rows-1))-1+columns]))
            ws8.cell(column=10,row=rows+11,value=('%s' %sk[rows-1]))
            ws8.cell(column=11,row=rows+11,value=('%s' %status[rows-1]))
    
    "Print nilai rata-rata per jam selama sebulan"
    for s in range(0,24):
        ws1.cell(column=s+2,row=37,value=('%s' %x_mean_hour_month[s]))        
        ws2.cell(column=s+2,row=37,value=('%s' %y_mean_hour_month[s]))
        ws3.cell(column=s+2,row=37,value=('%s' %z_mean_hour_month[s]))
        ws4.cell(column=s+2,row=37,value=('%s' %f_mean_hour_month[s]))
        ws5.cell(column=s+2,row=37,value=('%s' %h_mean_hour_month[s]))
        ws6.cell(column=s+2,row=37,value=('%s' %d_mean_hour_month[s]))
        ws7.cell(column=s+2,row=37,value=('%s' %i_mean_hour_month[s]))

    "Print total nilai sebulan"
    ws1['Z37'] = x_month
    ws2['Z37'] = y_month
    ws3['Z37'] = z_month
    ws4['Z37'] = f_month
    ws5['Z37'] = h_month
    ws6['Z37'] = d_month
    ws7['Z37'] = i_month

    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    thin_border_side = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style=None), 
                     bottom=Side(style=None))
    thin_border_updown = Border(left=Side(style=None), 
                     right=Side(style=None), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    hair_border = Border(left=Side(style='hair'), 
                     right=Side(style='hair'), 
                     top=Side(style='hair'), 
                     bottom=Side(style='hair'))
    hair_border_updown = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='hair'), 
                     bottom=Side(style='hair'))
    
    "FORMAT CELL XYZFHDI"
    "Setting cell data"
    for p in range(2,26):    
        for q in range(6,37):  
            ws1.cell(column=p,row=q).border = hair_border
            ws1.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws1.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws2.cell(column=p,row=q).border = hair_border
            ws2.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws2.cell(column=p,row=q).font = Font(size=9, name='Arial')            
            ws3.cell(column=p,row=q).border = hair_border
            ws3.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws3.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws4.cell(column=p,row=q).border = hair_border
            ws4.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws4.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws5.cell(column=p,row=q).border = hair_border
            ws5.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws5.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws6.cell(column=p,row=q).border = hair_border
            ws6.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws6.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws7.cell(column=p,row=q).border = hair_border
            ws7.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws7.cell(column=p,row=q).font = Font(size=9, name='Arial')
    "Setting cell Tanggal dan Mean per tanggal"
    for p in [1,26]:
         for r in range(4,38):
            ws1.cell(column=p,row=r).border = thin_border_side
            ws1.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws1.cell(column=p,row=r).font = Font(size=9, name='Arial') 
            ws2.cell(column=p,row=r).border = thin_border_side
            ws2.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws2.cell(column=p,row=r).font = Font(size=9, name='Arial')  
            ws3.cell(column=p,row=r).border = thin_border_side
            ws3.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws3.cell(column=p,row=r).font = Font(size=9, name='Arial')  
            ws4.cell(column=p,row=r).border = thin_border_side
            ws4.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws4.cell(column=p,row=r).font = Font(size=9, name='Arial')  
            ws5.cell(column=p,row=r).border = thin_border_side
            ws5.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws5.cell(column=p,row=r).font = Font(size=9, name='Arial')  
            ws6.cell(column=p,row=r).border = thin_border_side
            ws6.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws6.cell(column=p,row=r).font = Font(size=9, name='Arial')  
            ws7.cell(column=p,row=r).border = thin_border_side
            ws7.cell(column=p,row=r).alignment = Alignment(horizontal="center",vertical="center")
            ws7.cell(column=p,row=r).font = Font(size=9, name='Arial') 
    "Setting cell jam"
    for p in range(1,27): 
        for q in [4,5]:
            ws1.cell(column=p,row=q).border = thin_border
            ws1.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws1.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws2.cell(column=p,row=q).border = thin_border
            ws2.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws2.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws3.cell(column=p,row=q).border = thin_border
            ws3.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws3.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws4.cell(column=p,row=q).border = thin_border
            ws4.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws4.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws5.cell(column=p,row=q).border = thin_border
            ws5.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws5.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws6.cell(column=p,row=q).border = thin_border
            ws6.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws6.cell(column=p,row=q).font = Font(size=9, name='Arial')
            ws7.cell(column=p,row=q).border = thin_border
            ws7.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
            ws7.cell(column=p,row=q).font = Font(size=9, name='Arial')
        ws1.column_dimensions[get_column_letter(p)].width = 7
        ws2.column_dimensions[get_column_letter(p)].width = 7
        ws3.column_dimensions[get_column_letter(p)].width = 7
        ws4.column_dimensions[get_column_letter(p)].width = 7
        ws5.column_dimensions[get_column_letter(p)].width = 7
        ws6.column_dimensions[get_column_letter(p)].width = 7
        ws7.column_dimensions[get_column_letter(p)].width = 7
    
    "Setting cell mean per jam"
    for p in range(2,26):
        ws1.cell(column=p,row=37).border = thin_border_updown
        ws1.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws1.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws2.cell(column=p,row=37).border = thin_border_updown
        ws2.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws2.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws3.cell(column=p,row=37).border = thin_border_updown
        ws3.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws3.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws4.cell(column=p,row=37).border = thin_border_updown
        ws4.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws4.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws5.cell(column=p,row=37).border = thin_border_updown
        ws5.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws5.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws6.cell(column=p,row=37).border = thin_border_updown
        ws6.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws6.cell(column=p,row=37).font = Font(size=9, name='Arial')
        ws7.cell(column=p,row=37).border = thin_border_updown
        ws7.cell(column=p,row=37).alignment = Alignment(horizontal="center",vertical="center")
        ws7.cell(column=p,row=37).font = Font(size=9, name='Arial')
        
    for p in [1,26]:
        ws1.cell(column=p,row=37).border = thin_border
        ws2.cell(column=p,row=37).border = thin_border
        ws3.cell(column=p,row=37).border = thin_border
        ws4.cell(column=p,row=37).border = thin_border
        ws5.cell(column=p,row=37).border = thin_border
        ws6.cell(column=p,row=37).border = thin_border
        ws7.cell(column=p,row=37).border = thin_border
    
    "Format header"
    for p in[1,25]:
        ws1.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws2.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws3.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws4.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws5.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws6.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        ws7.cell(column=p,row=3).font = Font(size=10, bold=True, name='Arial')
        
    "Format judul"
    ws1.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws2.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws3.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws4.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws5.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws6.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')
    ws7.cell(column=1,row=2).font = Font(size=12, bold=True, name='Arial')        
    
    "SETTING CELL K INDEKS"
    "Format cell data"
    for q in range(12,43):
        ws8.cell(column=1,row=q).font = Font(size=11, bold=True, name='Arial')        
        for p in range(2,10):
            ws8.cell(column=p,row=q).border = hair_border
            ws8.cell(column=p,row=q).font = Font(size=11, name='Arial')        
        for p in[1,10,11]:
            ws8.cell(column=p,row=q).border = hair_border_updown
    for q in [10,11]:
        for p in range(1,12):
            ws8.cell(column=p,row=q).border = thin_border
            ws8.cell(column=p,row=q).font = Font(size=11, bold=True, name='Arial')        
    for p in range(1,12):
        ws8.cell(column=p,row=43).border = Border(top=Side(style='thin'))
    for p in range(1,11):
         ws8.column_dimensions[get_column_letter(p)].width = 7
    ws8.column_dimensions[get_column_letter(11)].width = 16
    for p in range(1,12):
        for q in range(8,43):
            ws8.cell(column=p,row=q).alignment = Alignment(horizontal="center",vertical="center")
    for p in [8,9]:
        for q in range(1,12):
            ws8.cell(column=q,row=p).border = thin_border
            ws8.cell(column=q,row=p).font = Font(size=11, bold=True, name='Arial')        
    for q in range(1,12):
        ws8.cell(column=q,row=1).font = Font(size=13, bold=True, name='Arial')  
    for p in range(1,43):
        ws8.row_dimensions[p].height = 16

    "SETTING CELL GEOMAGNETIC STORM"
    for q in range(1,16):
        for p in range(6,27):
            ws9.cell(column=q,row=p).border = hair_border_updown
        ws9.cell(column=q,row=27).border = Border(top=Side(style='thin'))
        for p in range(4,6):
            ws9.cell(column=q,row=p).border = thin_border
            ws9.cell(column=q,row=p).alignment = Alignment(horizontal="center",vertical="center")
            ws9.cell(column=q,row=p).font = Font(size=10, bold=True, name='Arial')
    for p in range(1,8):
        ws9.column_dimensions[get_column_letter(p)].width = 7
    for p in range(9,16):
        ws9.column_dimensions[get_column_letter(p)].width = 8
    ws9.column_dimensions[get_column_letter(8)].width = 17
    for p in range(1,16):
        for q in range(1,4):
            ws9.cell(column=p,row=q).font = Font(size=10, bold=True, name='Arial')
    for p in range(10,16):
        for q in range(2,4):
            ws9.cell(column=p,row=q).alignment = Alignment(horizontal="right",vertical="center")
            
    wb.save(filename = fileout)
    os.remove(pathIMFV + '\\data.dka')
    print '%s has been created' %(fileout)